from openpyxl import load_workbook
book = load_workbook('suppination.xlsm')
#print(book.sheetnames)
sheet= book.active

rows = sheet.rows
headers = [cell.value for cell in next(rows)]

all_rows= []

for row in rows:
    data = {}
    for title, cell in zip(headers, row):
        data[title]= cell.value
    all_rows.append(data)
print(all_rows)